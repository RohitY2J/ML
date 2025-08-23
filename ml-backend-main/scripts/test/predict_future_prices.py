import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split, TimeSeriesSplit
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.metrics import mean_squared_error, r2_score
import xgboost as xgb
from lightgbm import LGBMRegressor
from sklearn.neural_network import MLPRegressor
import joblib
from datetime import datetime, timedelta
import warnings
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings('ignore')

class PricePredictor:
    def __init__(self, target_days=5):
        self.target_days = target_days
        self.models = {
            'rf': RandomForestRegressor(n_estimators=200, max_depth=15, random_state=42),
            'gb': GradientBoostingRegressor(n_estimators=200, max_depth=5, random_state=42),
            'xgb': xgb.XGBRegressor(n_estimators=200, max_depth=5, random_state=42),
            'lgbm': LGBMRegressor(n_estimators=200, max_depth=5, random_state=42),
            'mlp': MLPRegressor(hidden_layer_sizes=(100, 50), max_iter=1000, random_state=42)
        }
        self.scalers = {}
        self.feature_importance = {}
        self.model_performance = {}

    def prepare_data(self, df):
        """Prepare data for prediction with confidence intervals"""
        # Sort by date and symbol
        df = df.sort_values(['symbol', 'date'])
        
        # Create future price targets
        for i in range(1, self.target_days + 1):
            df[f'future_price_{i}d'] = df.groupby('symbol')['close'].shift(-i)
        
        # Create features
        feature_columns = [
            'open', 'high', 'low', 'close', 'volume',
            'sma_20', 'sma_50', 'ema_20',
            'macd', 'macd_signal', 'macd_diff',
            'rsi_14', 'stoch_k', 'stoch_d',
            'bb_high', 'bb_low', 'bb_mid',
            'vwap', 'price_change', 'price_change_5d',
            'volume_change'
        ]
        
        # Add more sophisticated features
        df['price_volatility'] = df['high'] - df['low']
        df['price_momentum'] = df['close'].pct_change(periods=5)
        df['volume_momentum'] = df['volume'].pct_change(periods=5)
        df['bb_width'] = (df['bb_high'] - df['bb_low']) / df['bb_mid']
        df['rsi_momentum'] = df['rsi_14'].diff()
        
        feature_columns.extend([
            'price_volatility', 'price_momentum',
            'volume_momentum', 'bb_width', 'rsi_momentum'
        ])
        
        # Remove or replace inf/-inf and NaN values
        df = df.replace([np.inf, -np.inf], np.nan)
        df = df.dropna()
        
        return df, feature_columns

    def train_models(self, df, feature_columns):
        """Train multiple models with time series cross-validation"""
        for i in range(1, self.target_days + 1):
            target_column = f'future_price_{i}d'
            print(f"\nTraining models for {i}-day prediction...")
            
            # Prepare data
            X = df[feature_columns]
            y = df[target_column]
            
            # Time series split
            tscv = TimeSeriesSplit(n_splits=5)
            
            # Initialize storage for this prediction horizon
            self.scalers[i] = {}
            self.feature_importance[i] = {}
            self.model_performance[i] = {}
            
            # Train each model
            for model_name, model in self.models.items():
                print(f"\nTraining {model_name}...")
                
                # Store predictions for confidence interval
                all_predictions = []
                all_true_values = []
                
                # Time series cross-validation
                for train_idx, test_idx in tscv.split(X):
                    X_train, X_test = X.iloc[train_idx], X.iloc[test_idx]
                    y_train, y_test = y.iloc[train_idx], y.iloc[test_idx]
                    
                    # Scale features
                    scaler = StandardScaler()
                    X_train_scaled = scaler.fit_transform(X_train)
                    X_test_scaled = scaler.transform(X_test)
                    
                    # Train model
                    model.fit(X_train_scaled, y_train)
                    
                    # Make predictions
                    y_pred = model.predict(X_test_scaled)
                    all_predictions.extend(y_pred)
                    all_true_values.extend(y_test)
                
                # Calculate performance metrics
                mse = mean_squared_error(all_true_values, all_predictions)
                r2 = r2_score(all_true_values, all_predictions)
                
                # Store results
                self.scalers[i][model_name] = scaler
                self.model_performance[i][model_name] = {
                    'mse': mse,
                    'r2': r2,
                    'rmse': np.sqrt(mse)
                }
                
                # Store feature importance if available
                if hasattr(model, 'feature_importances_'):
                    self.feature_importance[i][model_name] = dict(
                        zip(feature_columns, model.feature_importances_)
                    )
                
                print(f"{model_name} - R²: {r2:.4f}, RMSE: {np.sqrt(mse):.2f}")

    def predict_with_confidence(self, symbol, current_data, feature_columns, confidence_level=0.95):
        """Make predictions with confidence intervals"""
        predictions = {}
        
        for i in range(1, self.target_days + 1):
            model_predictions = []
            
            # Get predictions from each model
            for model_name, model in self.models.items():
                scaler = self.scalers[i][model_name]
                X_scaled = scaler.transform(current_data[feature_columns])
                pred = model.predict(X_scaled)[0]
                model_predictions.append(pred)
            
            # Calculate ensemble prediction and confidence interval
            mean_prediction = np.mean(model_predictions)
            std_prediction = np.std(model_predictions)
            
            # Calculate confidence interval
            z_score = 1.96  # for 95% confidence interval
            lower_bound = mean_prediction - z_score * std_prediction
            upper_bound = mean_prediction + z_score * std_prediction
            
            predictions[i] = {
                'mean': mean_prediction,
                'lower': lower_bound,
                'upper': upper_bound,
                'std': std_prediction,
                'model_predictions': dict(zip(self.models.keys(), model_predictions))
            }
        
        return predictions

def save_predictions_to_excel(predictions_data, output_file='price_predictions.xlsx'):
    """Save predictions to Excel with formatting"""
    # Create Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Save main predictions
    predictions_data['main_predictions'].to_excel(writer, sheet_name='Predictions', index=False)
    
    # Save model performance
    predictions_data['model_performance'].to_excel(writer, sheet_name='Model Performance', index=False)
    
    # Get workbook and worksheet
    workbook = writer.book
    pred_sheet = workbook['Predictions']
    perf_sheet = workbook['Model Performance']
    
    # Define styles
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format Predictions sheet
    for col in range(1, pred_sheet.max_column + 1):
        # Format header
        cell = pred_sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Format data cells
        for row in range(2, pred_sheet.max_row + 1):
            cell = pred_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            # Format numbers
            if isinstance(cell.value, (int, float)):
                if 'price' in pred_sheet.cell(row=1, column=col).value.lower():
                    cell.number_format = '#,##0.00'
                elif 'change' in pred_sheet.cell(row=1, column=col).value.lower():
                    cell.number_format = '+0.00%;-0.00%'
    
    # Format Model Performance sheet
    for col in range(1, perf_sheet.max_column + 1):
        # Format header
        cell = perf_sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
        
        # Format data cells
        for row in range(2, perf_sheet.max_row + 1):
            cell = perf_sheet.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
            
            # Format numbers
            if isinstance(cell.value, (int, float)):
                if 'r2' in perf_sheet.cell(row=1, column=col).value.lower():
                    cell.number_format = '0.0000'
                else:
                    cell.number_format = '0.00'
    
    # Adjust column widths
    for sheet in [pred_sheet, perf_sheet]:
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    writer.close()
    print(f"\nPredictions saved to {output_file}")

def main():
    # Load the dataset
    print("Loading dataset...")
    df = pd.read_csv('ml_dataset.csv')
    
    # Initialize predictor
    predictor = PricePredictor(target_days=5)
    
    # Prepare data
    print("\nPreparing data...")
    df, feature_columns = predictor.prepare_data(df)
    
    # Train models
    print("\nTraining models...")
    predictor.train_models(df, feature_columns)
    
    # Save models and scalers
    print("\nSaving models...")
    import os
    os.makedirs('models', exist_ok=True)
    joblib.dump(predictor, 'models/price_predictor.joblib')
    
    # Make predictions for the most recent data
    print("\nMaking predictions...")
    latest_data = df.groupby('symbol').last().reset_index()
    
    # Prepare data for saving
    predictions_list = []
    model_performance_list = []
    
    for symbol in latest_data['symbol'].unique():
        symbol_data = latest_data[latest_data['symbol'] == symbol]
        current_price = symbol_data['close'].iloc[0]
        
        predictions = predictor.predict_with_confidence(
            symbol, symbol_data, feature_columns
        )
        
        # Store predictions
        for days, pred in predictions.items():
            predictions_list.append({
                'Symbol': symbol,
                'Current Price': current_price,
                'Prediction Days': days,
                'Expected Price': pred['mean'],
                'Price Change %': (pred['mean'] - current_price) / current_price,
                'Lower Bound': pred['lower'],
                'Upper Bound': pred['upper'],
                'Standard Deviation': pred['std']
            })
            
            # Store individual model predictions
            for model_name, model_pred in pred['model_predictions'].items():
                predictions_list[-1][f'{model_name} Prediction'] = model_pred
                predictions_list[-1][f'{model_name} Change %'] = (model_pred - current_price) / current_price
        
        # Store model performance
        for days, perf in predictor.model_performance.items():
            for model_name, metrics in perf.items():
                model_performance_list.append({
                    'Symbol': symbol,
                    'Prediction Days': days,
                    'Model': model_name,
                    'R² Score': metrics['r2'],
                    'RMSE': metrics['rmse']
                })
    
    # Create DataFrames
    predictions_df = pd.DataFrame(predictions_list)
    model_performance_df = pd.DataFrame(model_performance_list)
    
    # Save to CSV
    predictions_df.to_csv('price_predictions.csv', index=False)
    model_performance_df.to_csv('model_performance.csv', index=False)
    print("\nPredictions saved to price_predictions.csv")
    print("Model performance saved to model_performance.csv")
    
    # Save to Excel with formatting
    save_predictions_to_excel({
        'main_predictions': predictions_df,
        'model_performance': model_performance_df
    })
    
    # Print summary
    print("\nPrediction Summary:")
    print(predictions_df[['Symbol', 'Current Price', 'Prediction Days', 'Expected Price', 'Price Change %']].to_string(index=False))

if __name__ == "__main__":
    main() 