# Advanced Trading Signals Script
# Combines technical, zone, and ML-based signal generation
import pandas as pd
import numpy as np
import os
from pathlib import Path
import ta
from ta.trend import SMAIndicator, EMAIndicator, MACD
from ta.momentum import RSIIndicator, StochasticOscillator
from ta.volatility import BollingerBands
from ta.volume import VolumeWeightedAveragePrice
from sklearn.ensemble import RandomForestClassifier
from sklearn.preprocessing import StandardScaler
import psycopg2
from psycopg2.extras import execute_values
from tabulate import tabulate
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def get_db_connection():
    return psycopg2.connect(
        host="localhost",
        port="5433",
        database="stock_market",
        user="postgres",
        password="postgres"
    )

def load_and_prepare_data(data_dir='all-data'):
    all_data = []
    csv_files = list(Path(data_dir).glob('*_daily_data_*.csv'))
    for file_path in csv_files:
        try:
            df = pd.read_csv(file_path)
            df['date'] = pd.to_datetime(df['date'])
            df = df.sort_values('date')
            if 'symbol' not in df.columns:
                df['symbol'] = file_path.stem.split('_')[0]
            # Technical indicators
            df['sma_20'] = SMAIndicator(close=df['close'], window=20).sma_indicator()
            df['sma_50'] = SMAIndicator(close=df['close'], window=50).sma_indicator()
            df['ema_20'] = EMAIndicator(close=df['close'], window=20).ema_indicator()
            macd = MACD(close=df['close'])
            df['macd'] = macd.macd()
            df['macd_signal'] = macd.macd_signal()
            df['macd_diff'] = macd.macd_diff()
            df['rsi_14'] = RSIIndicator(close=df['close'], window=14).rsi()
            stoch = StochasticOscillator(high=df['high'], low=df['low'], close=df['close'])
            df['stoch_k'] = stoch.stoch()
            df['stoch_d'] = stoch.stoch_signal()
            bb = BollingerBands(close=df['close'])
            df['bb_high'] = bb.bollinger_hband()
            df['bb_low'] = bb.bollinger_lband()
            df['bb_mid'] = bb.bollinger_mavg()
            df['vwap'] = VolumeWeightedAveragePrice(
                high=df['high'], low=df['low'], close=df['close'], volume=df['volume']
            ).volume_weighted_average_price()
            df['price_change'] = df['close'].pct_change()
            df['price_change_5d'] = df['close'].pct_change(periods=5)
            df['volume_change'] = df['volume'].pct_change()
            df = df.dropna()
            all_data.append(df)
        except Exception as e:
            logger.warning(f"Error processing {file_path}: {str(e)}")
            continue
    if all_data:
        return pd.concat(all_data, ignore_index=True)
    return None

def get_trading_zones(conn, symbol, timeframe_days=90):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT zone_type, bottom_price, center_price, top_price
        FROM trading_zones
        WHERE symbol = %s AND timeframe_days = %s
    """, (symbol, timeframe_days))
    zones = cursor.fetchall()
    zone_dict = {}
    for zone in zones:
        zone_dict[zone[0]] = {
            'bottom': zone[1], 'center': zone[2], 'top': zone[3]
        }
    cursor.close()
    return zone_dict

def get_trendline(conn, symbol, timeframe_days=90):
    cursor = conn.cursor()
    cursor.execute("""
        SELECT trend_type, slope
        FROM trendlines
        WHERE symbol = %s AND timeframe_days = %s
        ORDER BY end_date DESC
        LIMIT 1
    """, (symbol, timeframe_days))
    result = cursor.fetchone()
    cursor.close()
    return result

def generate_signals(df, zones, trendline):
    last_row = df.iloc[-1]
    current_price = float(last_row['close'])
    signal = {'signal': 'HOLD', 'buy_target': None, 'sell_target': None, 'stop_loss': None}
    # Rule-based logic (combine best of both scripts)
    if zones:
        demand_zone = zones.get('immediate_demand_zone')
        supply_zone = zones.get('immediate_supply_zone')
        if demand_zone:
            demand_center = demand_zone['center']
            if abs((current_price - demand_center) / demand_center * 100) <= 2:
                signal['signal'] = 'BUY'
                signal['buy_target'] = demand_zone['bottom']
                signal['stop_loss'] = demand_zone['bottom'] * 0.98
        if supply_zone:
            supply_center = supply_zone['center']
            if abs((current_price - supply_center) / supply_center * 100) <= 2:
                signal['signal'] = 'SELL'
                signal['sell_target'] = supply_zone['top']
    # Technical indicator logic
    if last_row['sma_20'] > last_row['sma_50'] and current_price > last_row['sma_50']:
        signal['signal'] = 'BUY'
    if last_row['rsi_14'] > 70:
        signal['signal'] = 'SELL'
    if last_row['rsi_14'] < 30:
        signal['signal'] = 'BUY'
    if trendline:
        trend_type, slope = trendline
        if trend_type == 'uptrend' and slope > 0:
            signal['signal'] = 'BUY'
        elif trend_type == 'downtrend' and slope < 0:
            signal['signal'] = 'SELL'
    return signal

def train_ml_model(df):
    features = [
        'open', 'high', 'low', 'close', 'volume',
        'sma_20', 'sma_50', 'ema_20', 'macd', 'macd_signal', 'macd_diff',
        'rsi_14', 'stoch_k', 'stoch_d', 'bb_high', 'bb_low', 'bb_mid',
        'vwap', 'price_change', 'price_change_5d', 'volume_change'
    ]
    df = df.dropna()
    X = df[features]
    y = df['signal']
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X)
    model = RandomForestClassifier(n_estimators=100, random_state=42)
    model.fit(X_scaled, y)
    return model, scaler, features

def store_signals(conn, signals):
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS advanced_trading_signals (
            id SERIAL PRIMARY KEY,
            symbol VARCHAR(50),
            ltp DECIMAL(10,2),
            signal VARCHAR(10),
            buy_target DECIMAL(10,2),
            sell_target DECIMAL(10,2),
            stop_loss DECIMAL(10,2),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cursor.execute("DELETE FROM advanced_trading_signals")
    values = [(
        s['symbol'], s['ltp'], s['signal'], s['buy_target'], s['sell_target'], s['stop_loss']
    ) for s in signals]
    execute_values(cursor, """
        INSERT INTO advanced_trading_signals 
        (symbol, ltp, signal, buy_target, sell_target, stop_loss)
        VALUES %s
    """, values)
    conn.commit()
    cursor.close()

def export_signals(signals):
    df = pd.DataFrame(signals)
    df.to_csv('advanced_signals.csv', index=False)
    writer = pd.ExcelWriter('advanced_signals.xlsx', engine='openpyxl')
    df.to_excel(writer, sheet_name='Signals', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Signals']
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 15
    writer.close()

def display_signals(signals):
    df = pd.DataFrame(signals)
    print(tabulate(df, headers='keys', tablefmt='pretty', showindex=False))
    print(f"\nTotal signals: {len(signals)}")
    print(f"BUY: {sum(df['signal']=='BUY')}")
    print(f"SELL: {sum(df['signal']=='SELL')}")
    print(f"HOLD: {sum(df['signal']=='HOLD')}")

def main():
    df = load_and_prepare_data()
    if df is None:
        print("No data loaded.")
        return
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT symbol FROM trading_zones")
    symbols = [row[0] for row in cursor.fetchall()]
    signals = []
    model, scaler, features = train_ml_model(df)
    for symbol in symbols:
        df_symbol = df[df['symbol'] == symbol]
        if df_symbol.empty:
            continue
        zones = get_trading_zones(conn, symbol)
        trendline = get_trendline(conn, symbol)
        signal = generate_signals(df_symbol, zones, trendline)
        # ML prediction
        X_latest = df_symbol[features].iloc[[-1]]
        X_scaled = scaler.transform(X_latest)
        ml_pred = model.predict(X_scaled)[0]
        # Combine rule and ML
        final_signal = signal['signal']
        if final_signal == 'HOLD' and ml_pred == 1:
            final_signal = 'BUY'
        if final_signal == 'HOLD' and ml_pred == -1:
            final_signal = 'SELL'
        signals.append({
            'symbol': symbol,
            'ltp': float(df_symbol['close'].iloc[-1]),
            'signal': final_signal,
            'buy_target': signal.get('buy_target'),
            'sell_target': signal.get('sell_target'),
            'stop_loss': signal.get('stop_loss')
        })
    store_signals(conn, signals)
    export_signals(signals)
    display_signals(signals)
    conn.close()

if __name__ == "__main__":
    main() 