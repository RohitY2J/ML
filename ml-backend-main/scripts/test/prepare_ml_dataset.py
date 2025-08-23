import pandas as pd
import numpy as np
import os
from pathlib import Path
import ta
from ta.trend import SMAIndicator, EMAIndicator, MACD
from ta.momentum import RSIIndicator, StochasticOscillator
from ta.volatility import BollingerBands
from ta.volume import VolumeWeightedAveragePrice
from tabulate import tabulate
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import warnings

# Minimum required data points for reliable technical analysis
MIN_DATA_POINTS = 50  # Minimum number of data points required
MIN_DAYS = 30  # Minimum number of days required

def get_data_directory():
    """Get the absolute path to the data directory."""
    # Get the absolute path to the backend directory
    backend_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    # Return the path to the all-data directory
    return os.path.join(backend_dir, 'all-data')

def load_and_prepare_data(data_dir=None):
    """Load all CSV files and prepare the dataset with technical indicators."""
    all_data = []
    skipped_files = []
    
    # Use the correct data directory path
    if data_dir is None:
        data_dir = get_data_directory()
    
    print(f"Looking for CSV files in: {data_dir}")
    
    # Get all CSV files
    csv_files = list(Path(data_dir).glob('*_daily_data_*.csv'))
    
    if not csv_files:
        warnings.warn(f"No CSV files found in {data_dir} directory")
        return None
    
    print(f"Found {len(csv_files)} CSV files")
    
    for file_path in csv_files:
        try:
            # Read CSV file
            df = pd.read_csv(file_path)
            
            # Check if file has enough data points
            if len(df) < MIN_DATA_POINTS:
                warnings.warn(f"Insufficient data points in {file_path.name}: {len(df)} points (minimum {MIN_DATA_POINTS} required)")
                skipped_files.append((file_path.name, "insufficient_data_points"))
                continue
            
            # Convert date to datetime
            df['date'] = pd.to_datetime(df['date'])
            
            # Check if data spans enough days
            date_range = (df['date'].max() - df['date'].min()).days
            if date_range < MIN_DAYS:
                warnings.warn(f"Insufficient date range in {file_path.name}: {date_range} days (minimum {MIN_DAYS} days required)")
                skipped_files.append((file_path.name, "insufficient_date_range"))
                continue
            
            # Sort by date
            df = df.sort_values('date')
            
            # Add symbol column if not present
            if 'symbol' not in df.columns:
                df['symbol'] = file_path.stem.split('_')[0]
            
            # Check for missing values in essential columns
            essential_columns = ['open', 'high', 'low', 'close', 'volume']
            missing_values = df[essential_columns].isnull().sum()
            if missing_values.any():
                warnings.warn(f"Missing values found in {file_path.name}:\n{missing_values[missing_values > 0]}")
                skipped_files.append((file_path.name, "missing_values"))
                continue
            
            # Calculate technical indicators
            # Trend indicators
            df['sma_20'] = SMAIndicator(close=df['close'], window=20).sma_indicator()
            df['sma_50'] = SMAIndicator(close=df['close'], window=50).sma_indicator()
            df['ema_20'] = EMAIndicator(close=df['close'], window=20).ema_indicator()
            
            # MACD
            macd = MACD(close=df['close'])
            df['macd'] = macd.macd()
            df['macd_signal'] = macd.macd_signal()
            df['macd_diff'] = macd.macd_diff()
            
            # Momentum indicators
            df['rsi_14'] = RSIIndicator(close=df['close'], window=14).rsi()
            
            # Stochastic Oscillator
            stoch = StochasticOscillator(high=df['high'], low=df['low'], close=df['close'])
            df['stoch_k'] = stoch.stoch()
            df['stoch_d'] = stoch.stoch_signal()
            
            # Bollinger Bands
            bb = BollingerBands(close=df['close'])
            df['bb_high'] = bb.bollinger_hband()
            df['bb_low'] = bb.bollinger_lband()
            df['bb_mid'] = bb.bollinger_mavg()
            
            # Volume indicators
            df['vwap'] = VolumeWeightedAveragePrice(
                high=df['high'],
                low=df['low'],
                close=df['close'],
                volume=df['volume']
            ).volume_weighted_average_price()
            
            # Price changes
            df['price_change'] = df['close'].pct_change()
            df['price_change_5d'] = df['close'].pct_change(periods=5)
            
            # Volume changes
            df['volume_change'] = df['volume'].pct_change()
            
            # Generate target variable (BUY/SELL/HOLD signals)
            # Using a simple strategy: Buy when price crosses above SMA20, Sell when below
            df['signal'] = 0  # HOLD
            df.loc[df['close'] > df['sma_20'], 'signal'] = 1  # BUY
            df.loc[df['close'] < df['sma_20'], 'signal'] = -1  # SELL
            
            # Remove rows with NaN values (due to technical indicator calculations)
            df = df.dropna()
            
            # Check if enough data remains after technical indicator calculations
            if len(df) < MIN_DATA_POINTS:
                warnings.warn(f"Insufficient data points after technical calculations in {file_path.name}: {len(df)} points (minimum {MIN_DATA_POINTS} required)")
                skipped_files.append((file_path.name, "insufficient_data_after_calculations"))
                continue
            
            all_data.append(df)
            print(f"Successfully processed {file_path.name}")
            
        except Exception as e:
            warnings.warn(f"Error processing {file_path}: {str(e)}")
            skipped_files.append((file_path.name, str(e)))
            continue
    
    # Print summary of skipped files
    if skipped_files:
        print("\nSkipped Files Summary:")
        for filename, reason in skipped_files:
            print(f"- {filename}: {reason}")
    
    # Combine all dataframes
    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
        print(f"\nSuccessfully processed {len(all_data)} files")
        print(f"Skipped {len(skipped_files)} files")
        return combined_df
    else:
        warnings.warn("No data was successfully processed")
        return None

def format_excel(df, output_file='ml_dataset.xlsx'):
    """Save the dataset to an Excel file with formatting."""
    # Create a writer object
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Write the dataframe to Excel
    df.to_excel(writer, sheet_name='Trading Data', index=False)
    
    # Get the workbook and worksheet objects
    workbook = writer.book
    worksheet = writer.sheets['Trading Data']
    
    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Format headers
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        # Auto-adjust column width
        column_letter = get_column_letter(col)
        worksheet.column_dimensions[column_letter].width = 15
    
    # Format date column
    date_col = df.columns.get_loc('date') + 1
    worksheet.column_dimensions[get_column_letter(date_col)].width = 20
    
    # Format numeric columns
    numeric_cols = ['open', 'high', 'low', 'close', 'volume', 'sma_20', 'sma_50', 'ema_20',
                   'macd', 'macd_signal', 'macd_diff', 'rsi_14', 'stoch_k', 'stoch_d',
                   'bb_high', 'bb_low', 'bb_mid', 'vwap', 'price_change', 'price_change_5d',
                   'volume_change']
    
    for col_name in numeric_cols:
        if col_name in df.columns:
            col_idx = df.columns.get_loc(col_name) + 1
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = '#,##0.00'
    
    # Save the workbook
    writer.close()
    print(f"Excel file saved to {output_file}")

def display_sample(df, n_samples=5):
    """Display a sample of the data in a nicely formatted table."""
    # Select random samples
    sample_df = df.sample(n=min(n_samples, len(df)))
    
    # Format numeric columns
    numeric_cols = sample_df.select_dtypes(include=[np.number]).columns
    for col in numeric_cols:
        sample_df[col] = sample_df[col].round(2)
    
    # Convert to table
    table = tabulate(sample_df, headers='keys', tablefmt='pretty', showindex=False)
    print("\nSample of the dataset:")
    print(table)
    
    # Print dataset statistics
    print("\nDataset Statistics:")
    print(f"Total number of samples: {len(df)}")
    print(f"Number of features: {len(df.columns)}")
    print(f"Date range: {df['date'].min()} to {df['date'].max()}")
    print(f"Number of unique symbols: {df['symbol'].nunique()}")
    
    # Print signal distribution
    signal_dist = df['signal'].value_counts()
    print("\nSignal Distribution:")
    for signal, count in signal_dist.items():
        signal_name = {1: 'BUY', -1: 'SELL', 0: 'HOLD'}[signal]
        print(f"{signal_name}: {count} ({count/len(df)*100:.1f}%)")

def save_dataset(df, output_file='ml_dataset.csv'):
    """Save the prepared dataset to CSV and Excel files."""
    if df is not None:
        # Save to CSV
        df.to_csv(output_file, index=False)
        print(f"CSV file saved to {output_file}")
        
        # Save to Excel with formatting
        excel_file = output_file.replace('.csv', '.xlsx')
        format_excel(df, excel_file)
        
        # Display sample and statistics
        display_sample(df)
    else:
        print("No data to save")

if __name__ == "__main__":
    # Prepare the dataset
    df = load_and_prepare_data()
    
    # Save the dataset
    save_dataset(df) 